import openpyxl
from datetime import datetime
from Execute.executesql import get_connection
from io import BytesIO

TEMPLATE_PATH = "static/vehicle_checklist_template.xlsx"


def fetch_vehicle_from_db(n_sr_no):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            s_location_code,
            dt_entry_datetime,
            s_vehicle_no,
            s_vehicle_type,
            s_driver_name,
            s_contact_no,
            s_purpose_of_entry
        FROM dbo.VEHICLE_CHECK_LIST
        WHERE n_sr_no = ?
    """, (n_sr_no,))

    r = cursor.fetchone()
    cursor.close()
    conn.close()

    return {
        "s_location_code": r[0],
        "dt_entry_datetime": r[1],
        "s_vehicle_no": r[2],
        "s_vehicle_type": r[3],
        "s_driver_name": r[4],
        "s_contact_no": r[5],
        "s_purpose_of_entry": r[6]
    }


def generate_vehicle_checklist_excel(data):
    # If saved row → fetch from DB
    if "n_sr_no" in data:
        data = fetch_vehicle_from_db(data["n_sr_no"])

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ---- Fill cells (DO NOT CHANGE THESE) ----
    ws["C6"] = data.get("s_vehicle_no")
    ws["G6"] = data.get("s_vehicle_type")

    ws["C7"] = data.get("s_driver_name")
    ws["G7"] = data.get("s_contact_no")

    ws["C8"] = data.get("s_location_code")

    if data.get("dt_entry_datetime"):
        ws["G8"] = str(data.get("dt_entry_datetime"))[:10]

    ws["C9"] = data.get("s_purpose_of_entry")

    # ---- SAVE TO MEMORY (NOT FILE SYSTEM) ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Vehicle_Checklist_{data.get('s_vehicle_no','NA')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return output, filename
